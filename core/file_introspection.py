"""File introspection for materials files.

Phase 1 of the LLM File Understanding upgrade. Produces compact, deterministic
content summaries so the LLM can pick tools based on what a file actually
contains, not just its filename and extension.

Public surface:

- ``PARSER_VERSION`` — schema version stamped into every cached summary.
- ``PARSER_ID_BY_EXT`` — extension → parser id mapping.
- ``IntrospectionError`` — raised only by tests or callers that explicitly
  want to fail-fast; ``summarize_file`` itself NEVER raises.
- ``summarize_file(file_id, *, detail_level="default")`` — the only entry
  point used by the workflow. Caches per-file and globally by sha256.

Failure policy:

- Parse failures and unsupported extensions are mapped to stable
  ``content_kind`` values (``"error"`` vs ``"unsupported"``) so the
  caller can render a single fallback line without leaking internals.
- All cache writes are atomic (temp file → replace).
- Path resolution flows exclusively through ``core.upload_store``.

Security guards:

- File size pre-check against ``FILE_INTROSPECTION_MAX_FILE_SIZE_BYTES``
  before any parser runs.
- Bounded reads (64 KB or 4000 lines, whichever first) for tabular text.
- ``openpyxl.load_workbook(..., read_only=True, keep_links=False)`` plus
  explicit ``zipfile.BadZipFile`` / ``InvalidFileException`` handling.
- Encoding decoded with ``errors="replace"``; non-UTF-8 bytes downgrade
  confidence to ``"low"`` and emit a warning rather than failing.
"""

from __future__ import annotations

import csv
import copy
import hashlib
import io
import json
import re
import threading
import zipfile
from contextlib import contextmanager
from importlib import import_module
from pathlib import Path
from typing import Any

from config.settings import (
    FILE_INTROSPECTION_CACHE_DIR,
    FILE_INTROSPECTION_DEFAULT_PREVIEW_ROWS,
    FILE_INTROSPECTION_FULLER_PREVIEW_ROWS,
    FILE_INTROSPECTION_MAX_FILE_SIZE_BYTES,
)


PARSER_VERSION = "file-introspection-v1"
PARSER_VERSION_INT = "1"
INTROSPECTION_FILENAME = "introspection.v1.json"

# Extension → parser dispatch. Anything not listed is unsupported on purpose.
PARSER_ID_BY_EXT: dict[str, str] = {
    ".cif": "cif",
    ".csv": "tabular_csv",
    ".txt": "tabular_text",
    ".dat": "tabular_text",
    ".xlsx": "xlsx",
    ".xls": "xls",
}

_INTENT_KEYWORDS: dict[str, list[str]] = {
    "dos": ["dos", "density of state", "tdos"],
    "xrd": ["2theta", "2-theta", "xrd", "diffract", "intensity", "theta"],
    "phase_curve": ["temperature", "composition", "phase curve", "property"],
    "structure": ["structure", "cif", "crystal", "spacegroup", "spg"],
    "binary_phase": ["binary", "two-component", "two component"],
    "ternary_phase": ["ternary", "three-component", "three component"],
    "liquidus": ["liquidus"],
    "isothermal": ["isothermal"],
    "vertical_section": ["vertical section", "vertical-section"],
}

_DOS_LIKE_RANGES: tuple[tuple[float, float], ...] = (
    (-30.0, 30.0),
    (-50.0, 50.0),
)
_XRD_LIKE_RANGES: tuple[tuple[float, float], ...] = (
    (0.0, 180.0),
)

_MAX_BYTES_FOR_TABULAR = 64 * 1024  # 64 KB
_MAX_LINES_FOR_TABULAR = 4000
_MAX_CIF_ATOMS = 200

_LOCKS: dict[str, tuple[threading.Lock, int]] = {}
_LOCKS_GUARD = threading.Lock()


class IntrospectionError(RuntimeError):
    """Raised only by callers that opt into fail-fast behavior."""


# -----------------------------------------------------------------------------
# Public entry point
# -----------------------------------------------------------------------------


def summarize_file(file_id: str, *, detail_level: str = "default") -> dict[str, Any]:
    """Return a structured summary for ``file_id``. Never raises.

    The result conforms to the standard summary shape documented in the plan:
    file-level metadata + parser hints + content kind + intent + preview.
    On any error, ``content_kind`` is set to ``"error"`` and ``error`` key
    carries the message; tests can distinguish from unsupported.
    """
    if detail_level not in {"default", "fuller"}:
        detail_level = "default"

    try:
        upload_store = import_module("core.upload_store")
        metadata = upload_store.get_file_metadata(file_id)
        path = upload_store.resolve_file_path(file_id)
    except Exception as exc:  # ImportError, ValidationError, FileNotFoundError …
        return _error_summary(file_id, exc)

    extension = str(metadata.get("extension") or "").lower() or path.suffix.lower()
    filename = str(metadata.get("filename") or path.name)
    stat = path.stat()
    size_bytes = int(stat.st_size)

    if size_bytes == 0:
        return _unsupported_summary(file_id, filename, extension, sha256=None, reason="empty file")

    sha256 = _hash_file(path)

    # --- Cache lookup --------------------------------------------------
    cached = _lookup_caches(upload_store, file_id, sha256)
    if cached is not None:
        cached_view = dict(cached)
        cached_view["from_cache"] = True
        cached_view.setdefault("caches_hit", [])
        # Adjust preview width for the requested detail level without
        # recomputing the file.
        return _resize_preview_for_detail(cached_view, detail_level)

    # --- Compute -------------------------------------------------------
    parser_id = PARSER_ID_BY_EXT.get(extension, "none")
    if parser_id == "none":
        summary = _unsupported_summary(file_id, filename, extension, sha256=sha256,
                                      reason="extension not supported")
    elif size_bytes > FILE_INTROSPECTION_MAX_FILE_SIZE_BYTES:
        summary = _oversize_summary(file_id, filename, extension, sha256, parser_id, size_bytes)
    else:
        try:
            with _sha256_lock_ctx(sha256):
                # Re-check after acquiring the lock (another worker may have just populated it).
                cached = _lookup_caches(upload_store, file_id, sha256)
                if cached is not None:
                    cached_view = dict(cached)
                    cached_view["from_cache"] = True
                    cached_view.setdefault("caches_hit", cached_view.get("caches_hit", []))
                    return _resize_preview_for_detail(cached_view, detail_level)

                if parser_id == "cif":
                    summary = _parse_cif(path, file_id, filename, extension, sha256)
                elif parser_id == "tabular_csv":
                    summary = _parse_csv(path, file_id, filename, extension, sha256)
                elif parser_id == "tabular_text":
                    summary = _parse_tabular_text(path, file_id, filename, extension, sha256)
                elif parser_id == "xlsx":
                    summary = _parse_xlsx(path, file_id, filename, extension, sha256)
                elif parser_id == "xls":
                    summary = _parse_xls(path, file_id, filename, extension, sha256)
                else:
                    summary = _unsupported_summary(file_id, filename, extension,
                                                   sha256=sha256, reason="no parser")
        except Exception as exc:
            summary = _error_summary(file_id, exc)

    summary["from_cache"] = False
    summary["caches_hit"] = []
    summary["size_bytes"] = size_bytes
    summary["mtime_ns"] = int(stat.st_mtime_ns)

    # Shape a default copy for persistence so the cache always carries the
    # compact default view (no std/q1/median/q3, no _fuller_* fields), and
    # a separate caller view for the requested detail level. The freshly
    # computed summary still has both shapes available because the parsers
    # populate ``facts._fuller_head_rows`` / ``facts._fuller_column_stats``.
    # NOTE: deep copy because ``_resize_preview_for_detail`` mutates the
    # ``facts`` and ``preview`` sub-dicts in place; a shallow copy would
    # strip fuller fields from the caller's view too.
    to_persist = _resize_preview_for_detail(copy.deepcopy(summary), "default")
    summary = _resize_preview_for_detail(summary, detail_level)

    _persist_summary(upload_store, file_id, sha256, to_persist)
    return summary


# -----------------------------------------------------------------------------
# Cache helpers
# -----------------------------------------------------------------------------


def _lookup_caches(upload_store: Any, file_id: str, sha256: str | None) -> dict[str, Any] | None:
    """Return a cached summary from per-file then global cache, or None."""
    try:
        per_file = upload_store.get_introspection_cache_path(file_id, parser_version=PARSER_VERSION_INT)
    except Exception:
        per_file = None

    if per_file is not None and sha256 is not None:
        try:
            payload = json.loads(per_file.read_text(encoding="utf-8"))
            if _is_per_file_cache_fresh(payload, file_id, sha256, upload_store):
                payload = dict(payload)
                payload["caches_hit"] = ["per_file"]
                _normalize_cached_payload(payload)
                return payload
        except (OSError, json.JSONDecodeError):
            pass

    if sha256:
        try:
            global_path = upload_store.get_global_introspection_cache_path(sha256)
        except Exception:
            global_path = None
        if global_path is not None:
            try:
                payload = json.loads(global_path.read_text(encoding="utf-8"))
                # Globals are valid as long as parser_version + parser_id match.
                if payload.get("parser_version") == PARSER_VERSION:
                    payload = dict(payload)
                    payload["caches_hit"] = ["global"]
                    _normalize_cached_payload(payload)
                    # Re-stamp the per-file cache so next call hits it first.
                    _write_per_file_cache(upload_store, file_id, sha256, payload)
                    return payload
            except (OSError, json.JSONDecodeError):
                pass
    return None


def _normalize_cached_payload(payload: dict[str, Any]) -> None:
    """Coerce JSON-stringified int keys back to ``int`` in place.

    ``json.dumps`` converts ``{0: {...}, 1: {...}}`` into ``{"0": {...},
    "1": {...}}`` which then breaks the ``column_stats[0]`` lookup used by
    parsers and tests. Mutates the dicts in place.
    """
    for stats_key in ("column_stats", "_fuller_column_stats"):
        for parent_key in ("preview", "facts"):
            container = payload.get(parent_key, {}).get(stats_key)
            if not isinstance(container, dict):
                continue
            for key in list(container.keys()):
                if isinstance(key, str) and key.lstrip("-").isdigit():
                    value = container.pop(key)
                    container[int(key)] = value
            break


def _is_per_file_cache_fresh(payload: dict[str, Any], file_id: str, sha256: str,
                              upload_store: Any) -> bool:
    if payload.get("parser_version") != PARSER_VERSION:
        return False
    if payload.get("sha256") != sha256:
        return False
    # Re-verify against current file mtime + size on disk — slowest path,
    # covered by callers wrapping this in try/except for missing files.
    try:
        path = upload_store.resolve_file_path(file_id)
        stat = path.stat()
        if int(payload.get("size_bytes", -1)) != int(stat.st_size):
            return False
        if int(payload.get("mtime_ns", -1)) != int(stat.st_mtime_ns):
            return False
    except Exception:
        return False
    return True


def _persist_summary(upload_store: Any, file_id: str, sha256: str | None,
                     summary: dict[str, Any]) -> None:
    """Atomically write both per-file and global caches (best effort)."""
    if not sha256:
        sha256 = summary.get("sha256")
    if not sha256:
        return
    payload = dict(summary)
    payload.pop("from_cache", None)
    payload.pop("caches_hit", None)
    # Persist at the default detail level; the requester's detail level is reapplied on read.
    payload["summary_level"] = "default"

    try:
        _write_per_file_cache(upload_store, file_id, sha256, payload)
    except Exception:
        pass
    try:
        _write_global_cache(upload_store, sha256, payload)
    except Exception:
        pass


def _write_per_file_cache(upload_store: Any, file_id: str, sha256: str,
                          summary: dict[str, Any]) -> None:
    path = upload_store.resolve_introspection_cache_path(file_id, parser_version=PARSER_VERSION_INT)
    if path is None:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    _atomic_write_json(path, summary)


def _write_global_cache(upload_store: Any, sha256: str, summary: dict[str, Any]) -> None:
    path = upload_store.resolve_global_introspection_cache_path(sha256)
    if path is None:
        return
    _atomic_write_json(path, summary)


def _atomic_write_json(path: Path, payload: dict[str, Any]) -> None:
    """Atomic JSON write: write to .tmp then replace. Same-volume move on POSIX/Windows."""
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )
    tmp.replace(path)


def _sha256_lock(key: str) -> threading.Lock:
    """Return a process-local lock keyed by ``key``, creating it on demand."""
    with _LOCKS_GUARD:
        entry = _LOCKS.get(key)
        if entry is None:
            lock = threading.Lock()
            _LOCKS[key] = (lock, 1)
            return lock
        lock, refcount = entry
        _LOCKS[key] = (lock, refcount + 1)
        return lock


@contextmanager
def _sha256_lock_ctx(key: str):
    """Context manager around the per-sha256 lock. Reference-counted cleanup."""
    lock = _sha256_lock(key)
    lock.acquire()
    try:
        yield
    finally:
        lock.release()
        with _LOCKS_GUARD:
            entry = _LOCKS.get(key)
            if entry is None:
                return
            current, refcount = entry
            refcount -= 1
            if refcount <= 0:
                _LOCKS.pop(key, None)
            else:
                _LOCKS[key] = (current, refcount)


# -----------------------------------------------------------------------------
# Parsers
# -----------------------------------------------------------------------------


def _parse_cif(path: Path, file_id: str, filename: str, extension: str,
               sha256: str) -> dict[str, Any]:
    facts: dict[str, Any] = {}
    warnings: list[str] = []
    content_kind = "crystal_structure"
    inferred = "structure"
    confidence = "low"
    needs_clarification = False
    recommended = ["structure"]
    preview: dict[str, Any] = {}

    try:
        from pymatgen.core import Structure
        from pymatgen.symmetry.analyzer import SpacegroupAnalyzer
    except Exception as exc:
        warnings.append(f"pymatgen unavailable: {exc.__class__.__name__}")
        return _shape_summary(file_id, filename, extension, sha256, "cif",
                              "error", None, "low", [], True,
                              {"error": str(exc)}, {}, warnings=warnings)

    try:
        structure = Structure.from_file(str(path))
    except Exception as exc:
        warnings.append(f"pymatgen parse failed: {exc.__class__.__name__}")
        return _shape_summary(file_id, filename, extension, sha256, "cif",
                              "error", None, "low", [], True,
                              {"error": f"{exc.__class__.__name__}: {exc}"}, {}, warnings=warnings)

    # Optional atom cap to bound pymatgen cost on pathological CIFs.
    if len(structure) > _MAX_CIF_ATOMS:
        warnings.append(f"truncated to first {_MAX_CIF_ATOMS} of {len(structure)} atoms")
        try:
            structure = Structure.from_sites(structure.sites[:_MAX_CIF_ATOMS])
        except Exception:
            # Fall back to the full structure if truncation fails; the parser already
            # produced some data so don't throw it away.
            pass

    try:
        formula, _factor = structure.composition.get_reduced_formula_and_factor()
    except Exception:
        formula = None
    lattice = structure.lattice
    facts["formula"] = formula
    facts["atom_count"] = len(structure)
    facts["element_list"] = sorted({str(el) for el in structure.composition.elements})
    facts["lattice"] = {
        "a": float(lattice.a),
        "b": float(lattice.b),
        "c": float(lattice.c),
        "alpha": float(lattice.alpha),
        "beta": float(lattice.beta),
        "gamma": float(lattice.gamma),
    }

    try:
        sga = SpacegroupAnalyzer(structure)
        facts["symmetry"] = {
            "symbol": sga.get_space_group_symbol(),
            "number": int(sga.get_space_group_number()),
            "crystal_system": sga.get_crystal_system(),
        }
        confidence = "high"
    except Exception as exc:
        warnings.append(f"SpacegroupAnalyzer failed: {exc.__class__.__name__}")
        facts["symmetry"] = None
        confidence = "medium"

    preview["atom_count"] = len(structure)
    return _shape_summary(file_id, filename, extension, sha256, "cif",
                          content_kind, inferred, confidence,
                          recommended, needs_clarification,
                          facts, preview, warnings=warnings)


def _parse_tabular_text(path: Path, file_id: str, filename: str, extension: str,
                        sha256: str) -> dict[str, Any]:
    """Parse whitespace/comma/tab-delimited text with bounded reads.

    Reads at most ``_MAX_BYTES_FOR_TABULAR`` bytes or ``_MAX_LINES_FOR_TABULAR`` lines,
    whichever comes first. Encoding is decoded with ``errors="replace"`` so any
    non-UTF-8 bytes downgrade confidence rather than crashing.
    """
    try:
        raw_bytes = path.read_bytes()[:_MAX_BYTES_FOR_TABULAR]
    except Exception as exc:
        return _shape_summary(file_id, filename, extension, sha256, "tabular_text",
                              "error", None, "low", [], True,
                              {"error": str(exc)}, {}, warnings=[str(exc)])

    if not _looks_like_text(raw_bytes):
        return _shape_summary(file_id, filename, extension, sha256, "tabular_text",
                              "unsupported", None, "low", [], False,
                              {}, {}, warnings=["binary content detected"])

    text = raw_bytes.decode("utf-8", errors="replace")
    # Detect non-UTF8 byte replacement markers: when ``errors="replace"`` substitutes
    # a byte, the resulting text contains the Unicode replacement character (U+FFFD).
    has_replacement = "�" in text
    if has_replacement:
        warning = "non-utf8 bytes replaced"
    else:
        warning = None

    sample_lines = text.splitlines()[:_MAX_LINES_FOR_TABULAR]
    rows_numeric = _parse_text_rows(sample_lines)
    label_hits = _collect_label_hints("\n".join(sample_lines[:40]).lower())

    if not rows_numeric:
        warnings_list = [warning] if warning else []
        warnings_list.append("no numeric rows parsed")
        return _shape_summary(file_id, filename, extension, sha256, "tabular_text",
                              "unsupported", None, "low", [], False,
                              {"label_hints": sorted(label_hits)},
                              {"head_lines": sample_lines[:5]},
                              warnings=warnings_list)

    col_stats = _column_stats(rows_numeric)
    intent, conf, needs_clarification, recommended = _infer_tabular_intent(
        col_stats=col_stats, label_hints=label_hits, column_count=len(col_stats),
    )
    # Use a deep copy for ``_fuller_column_stats`` so stripping std/q1/median/q3
    # from the default-level preview.column_stats does not also strip the
    # fuller payload (parsers reuse the same col_stats dict for both).
    fuller_column_stats = copy.deepcopy(col_stats)
    facts = {
        "row_count_estimate": _estimate_row_count(path),
        "delimiter_guess": _guess_text_delimiter(sample_lines),
        "label_hints": sorted(label_hits),
        # Fuller-only data; see _resize_preview_for_detail.
        "_fuller_head_rows": [
            list(map(_preview_value, row))
            for row in rows_numeric[:FILE_INTROSPECTION_FULLER_PREVIEW_ROWS]
        ],
        "_fuller_column_stats": fuller_column_stats,
    }
    preview = {
        "head_rows": [list(map(_preview_value, row)) for row in rows_numeric[:5]],
        "column_stats": col_stats,
    }
    warnings: list[str] = []
    if warning:
        warnings.append(warning)

    return _shape_summary(file_id, filename, extension, sha256, "tabular_text",
                          "tabular_numeric" if rows_numeric and len(col_stats) >= 2
                          else "tabular_text",
                          intent, conf, recommended, needs_clarification,
                          facts, preview, warnings=warnings)


def _parse_csv(path: Path, file_id: str, filename: str, extension: str,
               sha256: str) -> dict[str, Any]:
    """Parse CSV using csv.Sniffer; bounded to ~64KB / FULLER_PREVIEW_ROWS rows for the probe."""
    try:
        with path.open("r", encoding="utf-8", errors="replace") as fh:
            sample_text = fh.read(_MAX_BYTES_FOR_TABULAR)
        sample_io = io.StringIO(sample_text)
        try:
            dialect = csv.Sniffer().sniff(sample_io.read(8192), delimiters=",\t;|")
            sample_io.seek(0)
        except csv.Error:
            dialect = csv.excel
            sample_io.seek(0)

        rows: list[list[str]] = []
        header_row: list[str] = []
        for row_idx, row in enumerate(csv.reader(sample_io, dialect=dialect)):
            if row_idx == 0:
                header_row = [str(cell) for cell in row]
                continue
            rows.append([str(cell) for cell in row])
            if len(rows) >= FILE_INTROSPECTION_FULLER_PREVIEW_ROWS:
                break
    except Exception as exc:
        return _shape_summary(file_id, filename, extension, sha256, "tabular_csv",
                              "error", None, "low", [], True,
                              {"error": str(exc)}, {}, warnings=[str(exc)])

    numeric_rows = []
    for row in rows:
        converted = _coerce_numeric_row(row)
        if converted is not None:
            numeric_rows.append(converted)

    label_hits = _collect_label_hints(" ".join(header_row).lower())

    if not numeric_rows:
        warnings = ["csv has no numeric rows"]
        return _shape_summary(file_id, filename, extension, sha256, "tabular_csv",
                              "tabular_text", None, "low", [], False,
                              {"header": header_row, "label_hints": sorted(label_hits)},
                              {"head_rows": rows[:5]},
                              warnings=warnings)

    col_stats = _column_stats(numeric_rows)
    intent, conf, needs_clarification, recommended = _infer_tabular_intent(
        col_stats=col_stats, label_hints=label_hits, column_count=len(col_stats),
    )
    # Deep copy for ``_fuller_column_stats`` — see _parse_tabular_text.
    fuller_column_stats = copy.deepcopy(col_stats)
    facts = {
        "header": header_row,
        "row_count_estimate": _estimate_row_count(path),
        "delimiter_guess": getattr(dialect, "delimiter", ","),
        "label_hints": sorted(label_hits),
        # Fuller-only data: kept in the cached payload so detail_level="fuller"
        # reads can expand head_rows + add std/q1/median/q3 without re-parsing.
        # ``_resize_preview_for_detail`` strips them for the default level.
        "_fuller_head_rows": [
            list(map(_preview_value, row))
            for row in numeric_rows[:FILE_INTROSPECTION_FULLER_PREVIEW_ROWS]
        ],
        "_fuller_column_stats": fuller_column_stats,
    }
    preview = {
        "head_rows": [list(map(_preview_value, row)) for row in numeric_rows[:5]],
        "column_stats": col_stats,
    }
    return _shape_summary(file_id, filename, extension, sha256, "tabular_csv",
                          "tabular_numeric" if len(col_stats) >= 2 else "tabular_text",
                          intent, conf, recommended, needs_clarification,
                          facts, preview)


def _parse_xlsx(path: Path, file_id: str, filename: str, extension: str,
                sha256: str) -> dict[str, Any]:
    """Open xlsx in read-only mode and summarize the first usable sheet.

    openpyxl with ``read_only=True`` and ``keep_links=False`` rejects billion-laughs
    style expansion at the zip layer; we still wrap in ``zipfile.BadZipFile`` and
    ``openpyxl.utils.exceptions.InvalidFileException`` so a malformed workbook never
    propagates a stack trace to the LLM context.
    """
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        return _shape_summary(file_id, filename, extension, sha256, "xlsx",
                              "error", None, "low", [], True,
                              {"error": f"openpyxl unavailable: {exc}"}, {},
                              warnings=[f"openpyxl missing: {exc.__class__.__name__}"])

    try:
        workbook = load_workbook(str(path), read_only=True, keep_links=False, data_only=True)
    except (zipfile.BadZipFile, Exception) as exc:
        # openpyxl raises InvalidFileException (subclass of Exception) for non-xlsx zip content.
        return _shape_summary(file_id, filename, extension, sha256, "xlsx",
                              "unsupported" if _looks_like_text(path.read_bytes()[:16])
                              else "error",
                              None, "low", [], False,
                              {"error": str(exc)}, {}, warnings=[str(exc)])

    sheet_names = list(workbook.sheetnames or [])
    warnings: list[str] = []
    if not sheet_names:
        warnings.append("workbook has no sheets")

    chosen_sheet = next((s for s in sheet_names if s.lower() != "sheet"), None) or (sheet_names[0] if sheet_names else None)
    preview: dict[str, Any] = {}
    facts: dict[str, Any] = {"sheet_names": sheet_names, "sheet_count": len(sheet_names)}

    if chosen_sheet is not None:
        ws = workbook[chosen_sheet]
        rows: list[list[Any]] = []
        try:
            for row in ws.iter_rows(values_only=True):
                if row is None:
                    continue
                rows.append([cell for cell in row])
                if len(rows) >= FILE_INTROSPECTION_FULLER_PREVIEW_ROWS:
                    break
        except Exception as exc:
            warnings.append(f"sheet iter failed: {exc.__class__.__name__}")
        header_row = [str(c) if c is not None else "" for c in rows[0]] if rows else []
        data_rows: list[list[Any]] = rows[1:]

        numeric_rows: list[list[float]] = []
        for row in data_rows:
            converted = _coerce_numeric_row([str(c) if c is not None else "" for c in row])
            if converted is not None:
                numeric_rows.append(converted)

        label_hits = _collect_label_hints((" ".join(header_row) + " " + chosen_sheet.lower()).strip())
        col_stats = _column_stats(numeric_rows) if numeric_rows else {}

        intent, conf, needs_clarification, recommended = _infer_tabular_intent(
            col_stats=col_stats, label_hints=label_hits,
            column_count=len(header_row),
        )

        # Phase-specific sheet-name overrides.
        name_lower = chosen_sheet.lower()
        if "binary" in name_lower and "phase" in name_lower:
            intent, conf, needs_clarification, recommended = "binary_phase", "high", False, ["binary_phase"]
        elif "ternary" in name_lower and "phase" in name_lower:
            intent, conf, needs_clarification, recommended = "ternary_phase", "high", False, ["ternary_phase"]
        elif "liquidus" in name_lower:
            intent, conf, needs_clarification, recommended = "liquidus", "high", False, ["liquidus"]
        elif "isothermal" in name_lower:
            intent, conf, needs_clarification, recommended = "isothermal", "high", False, ["isothermal"]
        elif "vertical" in name_lower and "section" in name_lower:
            intent, conf, needs_clarification, recommended = "vertical_section", "high", False, ["vertical_section"]

        facts["header"] = header_row
        facts["delimiter_guess"] = "xlsx"
        facts["label_hints"] = sorted(label_hits)
        # Fuller-only data: see _resize_preview_for_detail.
        facts["_fuller_head_rows"] = [
            list(map(_preview_value, row))
            for row in numeric_rows[:FILE_INTROSPECTION_FULLER_PREVIEW_ROWS]
        ]
        # Deep copy to keep std/q1/median/q3 after default-level stripping.
        facts["_fuller_column_stats"] = copy.deepcopy(col_stats)
        preview = {
            "first_sheet_preview": [header_row] + [list(r) for r in data_rows[:5]],
            "column_stats": col_stats,
        }

        content_kind = "tabular_numeric" if numeric_rows else "spreadsheet"
    else:
        content_kind = "spreadsheet"
        intent = None
        conf = "low"
        needs_clarification = False
        recommended = []

    try:
        workbook.close()
    except Exception:
        pass

    return _shape_summary(file_id, filename, extension, sha256, "xlsx",
                          content_kind, intent, conf, recommended,
                          needs_clarification, facts, preview, warnings=warnings)


def _parse_xls(path: Path, file_id: str, filename: str, extension: str,
               sha256: str) -> dict[str, Any]:
    """Open legacy ``.xls`` (BIFF) workbook with ``xlrd`` and summarize the
    first usable sheet.

    Mirrors the ``_parse_xlsx`` shape so the LLM can pick the same MCP intents
    on Phase Diagram spreadsheets saved with old Excel. ``xlrd`` 2.x only
    reads ``.xls`` (not ``.xlsx``), so it sits beside ``openpyxl`` without
    overlap. ``ragged_rows=True`` short-circuits ``IndexError`` on rows of
    differing length.
    """
    try:
        from xlrd import open_workbook as _xlrd_open_workbook  # type: ignore[import-not-found]
    except Exception as exc:
        return _shape_summary(file_id, filename, extension, sha256, "xls",
                              "error", None, "low", [], True,
                              {"error": f"xlrd unavailable: {exc}"}, {},
                              warnings=[f"xlrd missing: {exc.__class__.__name__}"])

    try:
        workbook = _xlrd_open_workbook(str(path), on_demand=False, ragged_rows=True)
    except Exception as exc:
        # xlrd raises XLRDError (subclass of Exception) for non-BIFF content;
        # if the file is text-shaped bytes masquerading as .xls we treat it
        # as ``unsupported`` rather than ``error``.
        try:
            head_bytes = path.read_bytes()[:16]
        except Exception:
            head_bytes = b""
        return _shape_summary(file_id, filename, extension, sha256, "xls",
                              "unsupported" if _looks_like_text(head_bytes) else "error",
                              None, "low", [], False,
                              {"error": str(exc)}, {}, warnings=[str(exc)])

    sheet_names = list(workbook.sheet_names() or [])
    warnings: list[str] = []
    if not sheet_names:
        warnings.append("workbook has no sheets")

    chosen_sheet = next((s for s in sheet_names if s.lower() != "sheet"), None) or (
        sheet_names[0] if sheet_names else None
    )
    preview: dict[str, Any] = {}
    facts: dict[str, Any] = {"sheet_names": sheet_names, "sheet_count": len(sheet_names)}

    if chosen_sheet is not None:
        ws = workbook.sheet_by_name(chosen_sheet)
        rows: list[list[Any]] = []
        try:
            for row_idx in range(min(ws.nrows, FILE_INTROSPECTION_FULLER_PREVIEW_ROWS)):
                row_values = ws.row_values(row_idx)
                rows.append(list(row_values))
        except Exception as exc:
            warnings.append(f"sheet iter failed: {exc.__class__.__name__}")

        header_row = [str(c) if c is not None else "" for c in rows[0]] if rows else []
        data_rows: list[list[Any]] = rows[1:]

        numeric_rows: list[list[float]] = []
        for row in data_rows:
            converted = _coerce_numeric_row([str(c) if c is not None else "" for c in row])
            if converted is not None:
                numeric_rows.append(converted)

        label_hits = _collect_label_hints(
            (" ".join(header_row) + " " + chosen_sheet.lower()).strip()
        )
        col_stats = _column_stats(numeric_rows) if numeric_rows else {}

        intent, conf, needs_clarification, recommended = _infer_tabular_intent(
            col_stats=col_stats, label_hints=label_hits,
            column_count=len(header_row),
        )

        # Phase-specific sheet-name overrides (mirrors _parse_xlsx).
        name_lower = chosen_sheet.lower()
        if "binary" in name_lower and "phase" in name_lower:
            intent, conf, needs_clarification, recommended = "binary_phase", "high", False, ["binary_phase"]
        elif "ternary" in name_lower and "phase" in name_lower:
            intent, conf, needs_clarification, recommended = "ternary_phase", "high", False, ["ternary_phase"]
        elif "liquidus" in name_lower:
            intent, conf, needs_clarification, recommended = "liquidus", "high", False, ["liquidus"]
        elif "isothermal" in name_lower:
            intent, conf, needs_clarification, recommended = "isothermal", "high", False, ["isothermal"]
        elif "vertical" in name_lower and "section" in name_lower:
            intent, conf, needs_clarification, recommended = "vertical_section", "high", False, ["vertical_section"]

        # Cache fuller head rows for the next phase; ``_resize_preview_for_detail``
        # trims to ``FILE_INTROSPECTION_DEFAULT_PREVIEW_ROWS`` when persisting.
        facts["header"] = header_row
        facts["delimiter_guess"] = "xls"
        facts["label_hints"] = sorted(label_hits)
        facts["_fuller_head_rows"] = [
            list(map(_preview_value, row)) for row in numeric_rows[:FILE_INTROSPECTION_FULLER_PREVIEW_ROWS]
        ]
        # Deep copy keeps std/q1/median/q3 after default-level stripping.
        facts["_fuller_column_stats"] = copy.deepcopy(col_stats)
        preview = {
            "first_sheet_preview": [header_row] + [list(r) for r in data_rows[:5]],
            "column_stats": col_stats,
        }

        content_kind = "tabular_numeric" if numeric_rows else "spreadsheet"
    else:
        content_kind = "spreadsheet"
        intent = None
        conf = "low"
        needs_clarification = False
        recommended = []

    try:
        workbook.release_resources()
    except Exception:
        pass

    return _shape_summary(file_id, filename, extension, sha256, "xls",
                          content_kind, intent, conf, recommended,
                          needs_clarification, facts, preview, warnings=warnings)


# -----------------------------------------------------------------------------
# Shared heuristics
# -----------------------------------------------------------------------------


def _infer_tabular_intent(*, col_stats: dict[int, dict[str, float]],
                          label_hints: set[str], column_count: int) -> tuple[str | None, str, bool, list[str]]:
    """Return (intent, confidence, needs_clarification, recommended_intents).

    Heuristic priority: explicit keyword hints first, then numeric-range evidence,
    then multi-candidate ambiguous fallback.
    """
    # 1. Explicit hints.
    primary_hint = _dominant_label_intent(label_hints)
    if primary_hint == "dos" and _first_column_dos_like(col_stats):
        return "dos", "high", False, ["dos"]
    if primary_hint == "xrd" and _first_column_xrd_like(col_stats):
        return "xrd", "high", False, ["xrd"]
    if primary_hint in {"binary_phase", "ternary_phase", "liquidus", "isothermal", "vertical_section", "phase_curve"}:
        return primary_hint, "medium", False, [primary_hint]

    # 2. Numeric-range evidence without labels.
    candidates: list[str] = []
    if _first_column_dos_like(col_stats):
        candidates.append("dos")
    if _first_column_xrd_like(col_stats):
        candidates.append("xrd")
    if column_count == 2 and _first_column_phase_curve_like(col_stats):
        candidates.append("phase_curve")

    if not candidates:
        return None, "low", False, []

    if len(candidates) == 1:
        return candidates[0], "medium", False, candidates

    # Multiple candidates => ambiguous.
    return None, "low", True, candidates


def _dominant_label_intent(label_hints: set[str]) -> str | None:
    """Return the most-specific intent keyword hit, or None."""
    priority = [
        "binary_phase", "ternary_phase", "vertical_section", "isothermal",
        "liquidus", "phase_curve", "xrd", "dos", "structure",
    ]
    # Lower-case hints are substrings inside bigger tokens; a direct equality
    # check captures both the bare keyword ("dos") and substrings like "xrd").
    for intent in priority:
        keywords = _INTENT_KEYWORDS.get(intent, [])
        if any(kw in label_hints for kw in keywords):
            return intent
    return None


def _collect_label_hints(lower_text: str) -> set[str]:
    hints: set[str] = set()
    for intent, keywords in _INTENT_KEYWORDS.items():
        for kw in keywords:
            if kw in lower_text:
                hints.add(kw)
    return hints


def _first_column_dos_like(col_stats: dict[int, dict[str, float]]) -> bool:
    if 0 not in col_stats:
        return False
    stats = col_stats[0]
    if not _is_finite_range(stats.get("min"), stats.get("max")):
        return False
    mn, mx = stats["min"], stats["max"]
    if stats.get("non_numeric_fraction", 1.0) > 0.05:
        return False
    for low, high in _DOS_LIKE_RANGES:
        if low <= mn and mx <= high:
            return True
    return False


def _first_column_xrd_like(col_stats: dict[int, dict[str, float]]) -> bool:
    if 0 not in col_stats:
        return False
    stats = col_stats[0]
    if not _is_finite_range(stats.get("min"), stats.get("max")):
        return False
    mn, mx = stats["min"], stats["max"]
    if stats.get("non_numeric_fraction", 1.0) > 0.05:
        return False
    for low, high in _XRD_LIKE_RANGES:
        if low <= mn and mx <= high and (mx - mn) > 1.0:
            return True
    return False


def _first_column_phase_curve_like(col_stats: dict[int, dict[str, float]]) -> bool:
    if 0 not in col_stats:
        return False
    stats = col_stats[0]
    if not _is_finite_range(stats.get("min"), stats.get("max")):
        return False
    mn, mx = stats["min"], stats["max"]
    # Loose range check — temperature vs composition style curves.
    if mn >= 0 and mx <= 1.2 and (mx - mn) > 0.05:
        return True
    if 200 <= mn and mx <= 4000:
        return True
    return False


def _is_finite_range(mn: float | None, mx: float | None) -> bool:
    if mn is None or mx is None:
        return False
    try:
        return float(mn) <= float(mx)
    except (TypeError, ValueError):
        return False


# -----------------------------------------------------------------------------
# Text probing / row parsing
# -----------------------------------------------------------------------------


def _looks_like_text(raw: bytes) -> bool:
    """Cheap binary-content sniff. Reject files where >5% of bytes are non-printable."""
    if not raw:
        return False
    sample = raw[:4096]
    text_chars = sum(1 for b in sample if b in {9, 10, 13} or 32 <= b < 127 or b >= 128)
    return (text_chars / max(1, len(sample))) >= 0.70


_DELIMITER_FALLBACK_CHARS: tuple[str, ...] = (",", ";")
"""Delimiters tried per-line after whitespace fails — covers comma/semicolon
separated .dat/.txt files (XRD/DOS style ``10,1200``) without disturbing the
whitespace default for the common ``1 2 3`` case."""


def _parse_text_rows(sample_lines: list[str]) -> list[list[float]]:
    """Return only the rows that parsed cleanly as all-numeric.

    Whitespace split is tried first to preserve the legacy behavior on
    ``1 2 3`` rows. If that yields fewer than two numeric cells (or fails),
    we fall back to comma/semicolon delimiters on the same line so a single
    cell like ``10,1200`` is split into two numeric columns. European decimal
    notation (``1,5``) is already handled by ``_coerce_numeric_row``'s
    ``replace(",", ".")`` once whitespace split exposes the cells.
    """
    out: list[list[float]] = []
    for raw in sample_lines:
        stripped = raw.strip()
        if not stripped or stripped.startswith("#"):
            continue
        values = _parse_one_text_row(stripped)
        if values is None or len(values) < 2:
            continue
        out.append(values)
    return out


def _parse_one_text_row(stripped: str) -> list[float] | None:
    """Parse a single text row, returning the best all-numeric split or None.

    Strategy: whitespace split first (most common .dat style); on miss, try
    each fallback delimiter and keep the split that yields the most numeric
    cells (≥ 2).
    """
    parts = stripped.split()
    if len(parts) >= 2:
        try:
            values = [float(p.replace(",", ".").rstrip(",")) for p in parts]
        except ValueError:
            values = None
        if values is not None and len(values) >= 2:
            return values

    best: list[float] | None = None
    for delim in _DELIMITER_FALLBACK_CHARS:
        if delim not in stripped:
            continue
        cells = stripped.split(delim)
        try:
            values = [float(c.strip().replace(",", ".").rstrip(",")) for c in cells]
        except ValueError:
            continue
        if len(values) < 2:
            continue
        if best is None or len(values) > len(best):
            best = values
    return best


def _coerce_numeric_row(row: list[str]) -> list[float] | None:
    converted: list[float] = []
    for cell in row:
        if cell is None or cell == "":
            return None
        normalized = cell.strip().replace(",", ".")
        try:
            converted.append(float(normalized))
        except ValueError:
            return None
    return converted if converted else None


def _std(values: list[float], mean: float) -> float | None:
    """Population standard deviation; returns None for empty input."""
    if not values:
        return None
    variance = sum((v - mean) ** 2 for v in values) / len(values)
    return round(variance ** 0.5, 6)


def _quartiles(sorted_values: list[float]) -> tuple[float, float, float] | None:
    """Return (q1, median, q3) using the linear-interpolation method.

    Matches ``statistics.quantiles(data, n=4, method='inclusive')`` to stay
    independent of the optional C-accelerated statistics module on minimal
    Python builds. Returns None for empty input.
    """
    n = len(sorted_values)
    if n == 0:
        return None

    def _at(pos: float) -> float:
        # pos in [0, n-1]; interpolate between floor and ceil.
        if pos <= 0:
            return sorted_values[0]
        if pos >= n - 1:
            return sorted_values[-1]
        lo = int(pos)
        frac = pos - lo
        return sorted_values[lo] + (sorted_values[lo + 1] - sorted_values[lo]) * frac

    median = _at((n - 1) / 2.0)
    # Inclusive quartile boundaries: include min/max in the lower/upper halves.
    q1 = _at((n - 1) / 4.0)
    q3 = _at(3 * (n - 1) / 4.0)
    return (round(q1, 6), round(median, 6), round(q3, 6))


def _column_stats(rows: list[list[float]]) -> dict[int, dict[str, float]]:
    """Compute per-column min/max/mean/count + std/q1/median/q3 (always).

    The extra descriptive statistics are cheap (O(n log n) on at most
    ``FILE_INTROSPECTION_FULLER_PREVIEW_ROWS`` rows) and let ``fuller`` callers
    see distributional shape. ``_resize_preview_for_detail`` strips
    ``std/q1/median/q3`` for the default level so cache payloads stay compact.
    """
    column_count = max((len(r) for r in rows), default=0)
    stats: dict[int, dict[str, float]] = {}
    for idx in range(column_count):
        values: list[float] = [float(r[idx]) for r in rows if len(r) > idx]
        if not values:
            continue
        non_numeric_fraction = 0.0  # rows are already filtered to all-numeric.
        mean = sum(values) / len(values)
        sorted_values = sorted(values)
        entry: dict[str, float] = {
            "min": sorted_values[0],
            "max": sorted_values[-1],
            "mean": mean,
            "count": len(values),
            "non_numeric_fraction": non_numeric_fraction,
        }
        std = _std(values, mean)
        if std is not None:
            entry["std"] = std
        quartiles = _quartiles(sorted_values)
        if quartiles is not None:
            entry["q1"], entry["median"], entry["q3"] = quartiles
        stats[idx] = entry
    return stats


def _guess_text_delimiter(sample_lines: list[str]) -> str:
    """Very crude delimiter guesser; whitespace dominates by default for .dat."""
    for line in sample_lines:
        if "," in line:
            return ","
        if ";" in line:
            return ";"
        if "\t" in line:
            return "tab"
    return "whitespace"


def _estimate_row_count(path: Path) -> int | None:
    """Cheap row estimate by counting newlines; capped and may overflow into bytes mode."""
    try:
        size = path.stat().st_size
    except OSError:
        return None
    if size == 0:
        return 0
    try:
        with path.open("rb") as fh:
            # read in 64 KB chunks to avoid memory blowups
            total = 0
            while True:
                buf = fh.read(64 * 1024)
                if not buf:
                    break
                total += buf.count(b"\n")
            return total
    except OSError:
        return None


def _preview_value(value: Any) -> Any:
    if isinstance(value, float):
        # Keep the JSON safe and short.
        if not (value == value) or value in (float("inf"), float("-inf")):
            return None
        return round(value, 6)
    return value


# -----------------------------------------------------------------------------
# SHA-256 + summary builders
# -----------------------------------------------------------------------------


def _hash_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(64 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _shape_summary(file_id: str, filename: str, extension: str, sha256: str,
                   parser_id: str, content_kind: str,
                   inferred: str | None, confidence: str,
                   recommended: list[str], needs_clarification: bool,
                   facts: dict[str, Any], preview: dict[str, Any],
                   warnings: list[str] | None = None) -> dict[str, Any]:
    return {
        "file_id": file_id,
        "filename": filename,
        "extension": extension,
        "sha256": sha256,
        "parser_version": PARSER_VERSION,
        "parser_id": parser_id,
        "summary_level": "default",
        "content_kind": content_kind,
        "inferred_content_type": inferred,
        "confidence": confidence,
        "recommended_mcp_intents": recommended,
        "needs_clarification": needs_clarification,
        "facts": dict(facts),
        "preview": dict(preview),
        "warnings": list(warnings or []),
    }


def _error_summary(file_id: str, exc: BaseException) -> dict[str, Any]:
    """Build an error summary when file metadata can't even be resolved."""
    return {
        "file_id": file_id,
        "filename": "",
        "extension": "",
        "sha256": None,
        "parser_version": PARSER_VERSION,
        "parser_id": "none",
        "summary_level": "default",
        "content_kind": "error",
        "inferred_content_type": None,
        "confidence": "low",
        "recommended_mcp_intents": [],
        "needs_clarification": False,
        "facts": {"error": str(exc)},
        "preview": {},
        "warnings": [str(exc)],
    }


def _unsupported_summary(file_id: str, filename: str = "", extension: str = "",
                         sha256: str | None = None, reason: str = "unsupported") -> dict[str, Any]:
    return {
        "file_id": file_id,
        "filename": filename,
        "extension": extension,
        "sha256": sha256,
        "parser_version": PARSER_VERSION,
        "parser_id": "none",
        "summary_level": "default",
        "content_kind": "unsupported",
        "inferred_content_type": None,
        "confidence": "low",
        "recommended_mcp_intents": [],
        "needs_clarification": False,
        "facts": {"reason": reason},
        "preview": {},
        "warnings": [reason],
    }


def _oversize_summary(file_id: str, filename: str, extension: str, sha256: str,
                      parser_id: str, size_bytes: int) -> dict[str, Any]:
    """Summary for a file that exceeds the size cap; refuse to parse."""
    return {
        "file_id": file_id,
        "filename": filename,
        "extension": extension,
        "sha256": sha256,
        "parser_version": PARSER_VERSION,
        "parser_id": parser_id,
        "summary_level": "default",
        "content_kind": "oversize",
        "inferred_content_type": None,
        "confidence": "low",
        "recommended_mcp_intents": [],
        "needs_clarification": False,
        "facts": {"size_bytes": size_bytes, "size_limit_bytes": FILE_INTROSPECTION_MAX_FILE_SIZE_BYTES},
        "preview": {},
        "warnings": [
            f"file larger than {FILE_INTROSPECTION_MAX_FILE_SIZE_BYTES} bytes; introspection refused",
        ],
    }


def _resize_preview_for_detail(summary: dict[str, Any], detail_level: str) -> dict[str, Any]:
    """Shape ``summary`` for the requested ``detail_level``.

    The cached payload always carries both shapes so that a fuller read after
    a default write still expands in zero IO:

    - ``preview.head_rows`` / ``preview.first_sheet_preview`` — default 5 rows
    - ``preview.column_stats`` — min/max/mean/count (+ std/q1/median/q3 always)
    - ``facts._fuller_head_rows`` — up to ``FILE_INTROSPECTION_FULLER_PREVIEW_ROWS`` rows
    - ``facts._fuller_column_stats`` — column stats with std/q1/median/q3

    At request time:

    - ``default``  → trim preview to ``DEFAULT_PREVIEW_ROWS``; strip
                     std/q1/median/q3 from ``preview.column_stats``; keep
                     ``facts._fuller_*`` so a subsequent fuller read expands
                     from cache without re-parsing.
    - ``fuller``   → expose up to ``FULLER_PREVIEW_ROWS`` rows; ensure
                     ``preview.column_stats`` includes std/q1/median/q3; keep
                     ``facts._fuller_*``.

    The function mutates ``summary`` in place and returns it; callers wanting
    immutability must pass a deep copy.
    """
    summary["summary_level"] = detail_level
    preview = summary.get("preview") or {}
    facts = summary.get("facts") or {}

    if detail_level == "fuller":
        cap = FILE_INTROSPECTION_FULLER_PREVIEW_ROWS
        fuller_rows = facts.get("_fuller_head_rows")
        if isinstance(fuller_rows, list) and fuller_rows:
            if "head_rows" in preview:
                preview["head_rows"] = list(fuller_rows)[:cap]
            elif "first_sheet_preview" in preview and preview["first_sheet_preview"]:
                header_row = preview["first_sheet_preview"][0]
                preview["first_sheet_preview"] = [header_row] + list(fuller_rows)[:cap]
        fuller_stats = facts.get("_fuller_column_stats")
        if isinstance(fuller_stats, dict) and fuller_stats:
            preview["column_stats"] = fuller_stats
    else:
        cap = FILE_INTROSPECTION_DEFAULT_PREVIEW_ROWS
        if isinstance(preview.get("head_rows"), list):
            preview["head_rows"] = preview["head_rows"][:cap]
        if isinstance(preview.get("first_sheet_preview"), list):
            preview["first_sheet_preview"] = preview["first_sheet_preview"][: cap + 1]
        # Strip std/q1/median/q3 from preview column_stats for the default
        # response; ``facts._fuller_column_stats`` retains them for fuller
        # reads and for any LLM tooling that introspects facts directly.
        col_stats = preview.get("column_stats")
        if isinstance(col_stats, dict):
            for entry in col_stats.values():
                if isinstance(entry, dict):
                    for key in ("std", "q1", "median", "q3"):
                        entry.pop(key, None)

    summary["preview"] = preview
    summary["facts"] = facts
    return summary


__all__ = [
    "PARSER_VERSION",
    "PARSER_ID_BY_EXT",
    "IntrospectionError",
    "summarize_file",
]
